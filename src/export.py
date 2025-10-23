try:
    from openpyxl import Workbook
    from openpyxl.utils import get_column_letter
    from openpyxl.styles import Font
except ImportError:
    raise SystemExit("openpyxl not installed. Run:  pip install openpyxl")

from db import connect, _fetch_items_for_export, _fetch_ratings_ledger

def _bucket_by_media(items, order=None, include_empty=False):
    """
    items: list of dict rows from _fetch_items_for_export(...)
    order: optional explicit order like ["game","book","movie","tv","music"]
    include_empty: if True, create empty sheets for media in 'order' even when no rows
    returns: Ordered dict-like {sheet_title: {"rows": [...]}}
    """
    # friendly sheet names (<=31 chars for Excel)
    DISPLAY = {
        "game": "Games",
        "book": "Books",
        "movie": "Movies",
        "tv": "TV",
        "music": "Music",
    }
    # group rows
    grouped = {}
    for r in items:
        code = (r.get("media_code") or "").lower()
        grouped.setdefault(code, []).append(r)

    # determine order
    if order:
        codes = order
    else:
        # keep a natural, friendly default order
        default = ["game","book","movie","tv","music"]
        present = [c for c in default if c in grouped]
        # add any unknown codes at the end
        others = [c for c in grouped.keys() if c not in present]
        codes = present + others

    sheets = {}
    for code in codes:
        rows = grouped.get(code, [])
        if rows or include_empty:
            title = DISPLAY.get(code, (code or "Other")).strip()[:31]
            sheets[title] = {"rows": rows}
    return sheets

def cmd_export_xlsx(args):
    conn = connect(args.db)

    # Pull items (optionally filtered by a single media/platform)
    items = _fetch_items_for_export(
        conn,
        media=args.media,              # note: if you set --split-by-media, consider leaving this None
        platform=args.platform,
        min_itchio=args.min_itchio,
        limit=args.limit
    )

    sheets = {}
    if args.split_by_media:
        # Optional explicit order via --tab-order=game,book,...
        order = [s.strip() for s in (args.tab_order or "").split(",") if s.strip()] or None
        sheets.update(_bucket_by_media(items, order=order, include_empty=args.include_empty_tabs))
    else:
        # Single 'Items' sheet (current behavior)
        sheets["Items"] = {"rows": items}

    # Optional Ratings sheet (unchanged behavior)
    if args.include_ratings:
        ratings = _fetch_ratings_ledger(
            conn,
            media=args.media if not args.split_by_media else None,  # if split, include all ratings
            source=args.source,
            since=args.since,
            limit=args.limit_ratings
        )
        sheets["Ratings"] = {"rows": ratings}

    # Write the workbook
    _write_xlsx(args.out, sheets)
    # Friendly summary
    tab_counts = ", ".join(f"{name}:{len(payload['rows'])}" for name, payload in sheets.items())
    print(f"wrote Excel → {args.out}  ({tab_counts})")

def _write_xlsx(path, sheets_dict):
    """
    sheets_dict = {
      "Items": {"rows": [ {col:val,...}, ... ]},
      "Ratings": {"rows": [...]}
    }
    """
    
    wb = Workbook()
    # Remove default sheet
    wb.remove(wb.active)

    for sheet_name, payload in sheets_dict.items():
        ws = wb.create_sheet(title=sheet_name[:31])  # Excel sheet name limit
        rows = payload.get("rows", [])
        # headers
        if rows:
            headers = list(rows[0].keys())
        else:
            headers = ["(no data)"]
        ws.append(headers)
        for cell in ws[1]:
            cell.font = Font(bold=True)
        # data
        for r in rows:
            ws.append([r.get(h) for h in headers])
        # autosize columns
        for col_idx, header in enumerate(headers, start=1):
            max_len = max((len(str(header)),) + tuple(len(str(r.get(header))) for r in rows)) if rows else len(str(header))
            ws.column_dimensions[get_column_letter(col_idx)].width = min(max_len + 2, 60)

    wb.save(path)
